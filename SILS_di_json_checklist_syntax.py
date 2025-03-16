# import libraries 
from datetime import datetime
import json
import os 
import random
import shutil
import sys
from io import BytesIO

from concurrent.futures import ThreadPoolExecutor, as_completed

from langchain.output_parsers.list import NumberedListOutputParser
from langchain.prompts.prompt import PromptTemplate

from langchain_community.llms import Ollama

from openpyxl import load_workbook, Workbook

# from langchain.chains.llm import LLMChain
from langchain_community.llms import Ollama
from langchain.prompts import PromptTemplate

from typing import Optional

# LLM model
model_select = 'llama3.1'

random.seed(2024)

format = sys.argv[3]
worksheet = load_workbook(sys.argv[2])[format]

output_report_container_name =  "quesnel-ground"

def create_prompt(format_instructions):
    QA_TEMPLATE = """
        Check the VALUE, surrounded by ``` below, if they follow the DESCRIPTION below, and generally similar to the EXAMPLE. 
        If EXAMPLE is Nil, then any VALUE is acceptable. 
        VALUE: ```{value}```
        DESCRIPTION: ```{description}```
        EXAMPLE: ```{example}```
        If it follows, then return just one of these two answer: True or False.
        The output should be just a boolean value.
        Answer: {format_instructions}
    """
    #Verify your answer, and if the result list has more than 2 items, then Value has multiple parts. Treat them all as one value only, and ignore the number in brackets in them. Retry to shorten it to format above.
    return PromptTemplate(
        input_variables=["value", "description", "example"], 
        partial_variables={"format_instructions": format_instructions},
        template=QA_TEMPLATE)

def llm_check_value(text_content, definition, example):
    output_parser = NumberedListOutputParser()  
    format_instructions = output_parser.get_format_instructions()

    llm = Ollama(model = model_select, temperature = 0.0)
    prompt = create_prompt(format_instructions)

    llm_chain = prompt | llm | output_parser
    # LLM_start_time = datetime.now()
    result = llm_chain.invoke({'value': text_content, 'description': definition, 'example': example})
    # print(result)
    # return result
    if eval(result[0]):
        return [result, 'No explanation']
    else:
        attempt = 0
        while not result or len(result) != 2:
            result = llm_chain.invoke({'value': text_content, 'instruction': definition, 'example': example})
                # print("-", end = "", flush = True)
            if attempt > 4:
                result = ['False', 'Too many attempts']
                # print('\t', result[1])
                break
            else:
                attempt = attempt + 1
    # print("--- QA Result:", result, '---', end = ' ')
    return result

def load_worksheet_as_dict():
    result_dict = {}
    for row_i, row in enumerate(worksheet['D'], 1):
        # print(row)
        # print(worksheet.cell(row_i, column = 4).value)
        if row.value is not None:
            result_dict[row.value.lower().replace(':', '')] = row_i
    return result_dict

worksheet_dict = load_worksheet_as_dict()
# print(worksheet_dict)

def check_field(field_name):
    # print(field_name.replace(':', ''), end = '\t')
    row_i = worksheet_dict.get(field_name.lower().replace(':', ''))
    # print(row_i)
    if row_i is None:
        return ['Not found field','NA', 'NA']
    return [
        worksheet.cell(row=row_i, column=6).value, 
        worksheet.cell(row=row_i, column=7).value
    ]

def clean_value(value):
    to_be_clean = ['\n', '*', ':']
    for char in to_be_clean:
        value = value.replace(char, '')
    print(value)
    return value

def process_json(json_file):
    # start_time = datetime.now()
    error_list = []
    error_counter = 0
    json_error_dict = {}
    for field, value in json_file.items():
        print(f"{field}\t:\t{value}", flush = True)
        if type(value[0]) is not str:
            print(value, '\n\t-', value[0], 'is not a string')
            continue
        else: 
            definition, example = check_field(field)
            # print(value_format, definition, example)
            # print('111', flush = True)
            content = clean_value(value[0])
            if definition == 'NA' or content is None:
                print('content', content)
                continue
            else:
                check, standard = llm_check_value(content, definition, example)
                # print('222', flush=True)
                # print(check)
                if eval(check[0]):
                    # print(".", end = '', flush = True)
                    # print(f"{content} PASSED\n", flush = True) #:\t{content}
                    error_list = []
                    json_error_dict[field] = error_list
                    # print(json_error_dict)
                    # continue
                else:
                    error_counter += 1
                    error_list.append(standard)
                    json_error_dict[field] = error_list
                    # print(json_error_dict)
                    # print(f"<{content}> FAILED\n\t{standard}\n", flush = True)
    # end_time = datetime.now()
    json_error_dict['error_score'] = error_counter
    print(f"\nQA Completed - Error score: {error_counter}", flush = True)
    return json_error_dict

# # function writes to Azure, push data
# def write_to_azure(file_name, output_file, container_name):
#     blob_service_client = BlobServiceClient.from_connection_string(output_connection_string)
#     container_client = blob_service_client.get_container_client(container_name)
#     with open(output_file, "rb") as data:
#             container_client.upload_blob(name=file_name, data=data, overwrite=True)

def write_to_file(file_list_in_directory, error_dict):
    for json_file in file_list_in_directory:
        if json_file.endswith('.json'):
            json_uuid = json_file[:-5]
            file_address = os.path.join(sys.argv[1], json_file)
            if json_uuid in error_dict:
                json_data = error_dict[json_uuid]
            else:
                print(f"{json_uuid} not found in error_dict", flush = True)
                continue
            if json_data['error_score'] != 0:
                destination_file = './'+ format + '_syntax_error' + '/' + json_file
                if not os.path.isdir('./'+ format + '_syntax_error' + '/'):
                    os.mkdir('./'+ format + '_syntax_error' )
                print(f"{json_uuid} FAILED", flush = True, end = '')
                shutil.copy(file_address, destination_file)
            else: 
                destination_file = './'+ format + '_success_json' + '/' + json_file
                if not os.path.isdir('./'+ format + '_success_json' + '/'):
                    os.mkdir('./'+ format + '_success_json' )
                print(f"{json_uuid} passed", flush = True, end = '')
                shutil.copy(file_address, destination_file)
    report_name = str(format) + '_report.json'
    json_error_data = json.dumps(error_dict, indent=4)
    with open(report_name, 'w') as report_file:
        report_file.write(json_error_data)

def check_json_syntax_local():
    start_time = datetime.now()
    file_list_in_directory = os.listdir(sys.argv[1])
    loaded_list_of_json_files = {}
    for json_file in file_list_in_directory:
        if json_file.endswith('.json'):
            file_address = os.path.join(sys.argv[1], json_file)
            uuid = json_file[:-5]
            # print(uuid)
            with open(file_address, 'rt', encoding='utf-8') as file:
                doc = json.load(file)
                loaded_list_of_json_files[uuid] = doc
        else:
            continue
    error_dict = {}
    for uuid, json_file in loaded_list_of_json_files.items():
        print(uuid)
        error_dict[uuid] = {'error_score': 1}
        try:
            error_dict[uuid] = process_json(json_file)
            ### error_dict[2019.json]: {'Date': [], 'error_score' = 0}###
            # print(error_dict)
        except:
            print('error hit for', uuid)
            continue
    # print(error_dict)
    write_to_file(file_list_in_directory, error_dict)

    end_time = datetime.now()
    seconds = (end_time - start_time).total_seconds()
    print(f"Total Execution time: {seconds} secs for {len(file_list_in_directory)} files at {end_time}", flush=True)
        
###---------------------------------------------------------------###
if __name__ == "__main__": 
    print("Running local at", datetime.now())
    check_json_syntax_local()

## Example: python .\python\di_json_checklist_syntax.py .\json\sils-ground\raw_jsons\ .\Data_Standards\Data_Standard_Quesnel_Roving.xlsx 'Data_Standard_Quesnel_Roving'